#!/usr/bin/env python3
"""Mutation Aggregator v6 (Clean Evidence)
Ne garde que les données pertinentes avec preuve clinique/pharmacologique.
Sources: UniProt (Base), ClinVar (Clinical), PharmGKB (Drug Proof).
"""
from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import sqlite3
import threading
import time
from abc import ABC, abstractmethod
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime, UTC
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote_plus

import pandas as pd
import requests
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from tqdm import tqdm

# --- Configuration ---
CONFIG_PATH = os.environ.get("MUT_AGG_CONFIG", "config.yaml")
CACHE_DB = os.environ.get("MUT_AGG_CACHE", "mutation_cache.sqlite3")
LOG_FILE = os.environ.get("MUT_AGG_LOG", "mutation_aggregator.log")
DEFAULT_BATCH_FILE = "uniprot_ids.csv"

# =============================================================================
# 1. LOGGING
# =============================================================================

class ContextualFilter(logging.Filter):
    def filter(self, record): record.uid = getattr(record, "uid", "GLOBAL"); return True

def setup_logging(level: str = "INFO") -> logging.Logger:
    formatter = logging.Formatter("[%(asctime)s] [%(levelname)s] [%(uid)s] %(message)s", datefmt="%H:%M:%S")
    stream_h = logging.StreamHandler()
    stream_h.setFormatter(formatter)
    stream_h.setLevel(level)
    file_h = logging.FileHandler(LOG_FILE)
    file_h.setFormatter(formatter)
    file_h.setLevel(logging.DEBUG)
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.DEBUG)
    root_logger.handlers.clear()
    root_logger.addHandler(stream_h)
    root_logger.addHandler(file_h)
    for h in root_logger.handlers: h.addFilter(ContextualFilter())
    return root_logger

logger = setup_logging()

# =============================================================================
# 2. PARSING (Simple & Robuste)
# =============================================================================

class VariantParser:
    """Parse les variants protéiques ET génomiques."""
    AA_3TO1 = { "Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D", "Cys": "C", "Gln": "Q", "Glu": "E",
                "Gly": "G", "His": "H", "Ile": "I", "Leu": "L", "Lys": "K", "Met": "M", "Phe": "F",
                "Pro": "P", "Ser": "S", "Thr": "T", "Trp": "W", "Tyr": "Y", "Val": "V" }
    
    # Protein patterns
    RE_HGVS_3 = re.compile(r"p\.([A-Za-z]{3})(\d+)([A-Za-z]{3}|[=*?])")
    RE_SIMPLE = re.compile(r"(?<![A-Za-z])([A-Z])(\d+)([A-Z*=])(?![A-Za-z0-9])")
    
    # Genomic patterns
    RE_RSID = re.compile(r"rs(\d+)")
    RE_INTRON = re.compile(r"c\.(-?\d+)([+-])(\d+)([ACGT]>[ACGT])?")
    RE_CHR_POS = re.compile(r"Chr(\d+|X|Y):(\d+)")

    VALID_AA_1 = set("ACDEFGHIKLMNPQRSTVWY*")
    GENE_PATTERNS = re.compile(r"(SLC|ABC|CYP|UGT|NAT|TPMT|DPYD|VKORC|OATP|OAT|OCT)\w*", re.I)
    
    # Configuration splicing (surchargeable via set_config)
    _splice_config: Dict = {}
    
    @classmethod
    def set_config(cls, config: Dict):
        """Configure les seuils de splicing depuis la config."""
        cls._splice_config = config.get("splicing", {})

    @classmethod
    def normalize_hgvs_c(cls, hgvs: str) -> str:
        """Normalise hgvs_c en extrayant juste la partie c.XXX."""
        if not hgvs:
            return ""
        # Format: NM_006446.5(SLCO1B1):c.388A>G -> c.388A>G
        if ":c." in hgvs:
            return "c." + hgvs.split(":c.")[-1]
        if hgvs.startswith("c."):
            return hgvs
        return hgvs

    @classmethod
    def parse_protein(cls, raw: Any) -> Tuple[Optional[int], Optional[str], Optional[str]]:
        """Parse la position protéique avec validation stricte."""
        if not raw: return None, None, None
        text = str(raw)
        
        # Priorité 1: HGVS 3-lettre (p.Asn130Asp) - le plus fiable
        if m := cls.RE_HGVS_3.search(text):
            ref = cls.AA_3TO1.get(m[1].capitalize(), m[1])
            pos = int(m[2])
            alt_raw = m[3]
            if alt_raw in ("=", "?"):
                alt = None
            elif alt_raw == "*":
                alt = "*"
            else:
                alt = cls.AA_3TO1.get(alt_raw.capitalize(), alt_raw)
            # Valider les acides aminés
            if ref not in cls.VALID_AA_1 or (alt and alt not in cls.VALID_AA_1):
                return None, None, None
            return pos, ref, alt
        
        # Priorité 2: Format simple (M235T) - avec validation contextuelle
        if m := cls.RE_SIMPLE.search(text):
            # Vérifier contexte de nom de gène
            match_text = text[max(0, m.start()-10):m.end()+10]
            if cls.GENE_PATTERNS.search(match_text):
                # Vérifier si le match est isolé
                if m.start() > 0 and text[m.start()-1].isalpha():
                    return None, None, None
                if m.end() < len(text) and text[m.end()].isalnum():
                    return None, None, None
            
            ref, alt = m[1], m[3] if m[3] != "=" else None
            if ref not in cls.VALID_AA_1 or (alt and alt not in cls.VALID_AA_1):
                return None, None, None
            return int(m[2]), ref, alt
        
        return None, None, None

    @classmethod
    def parse_rsid(cls, raw: Any) -> Optional[str]:
        """Extrait le rsID."""
        if not raw: return None
        if m := cls.RE_RSID.search(str(raw)):
            return f"rs{m[1]}"
        return None

    @classmethod
    def parse_genomic_location(cls, raw: Any) -> Dict[str, Any]:
        """Extrait la position génomique et intronique."""
        if not raw: return {}
        text = str(raw)
        result = {}
        
        # Intron parsing: c.1234+5G>A ou c.1234-10A>T
        if m := cls.RE_INTRON.search(text):
            exon_pos = int(m[1])
            direction = m[2]  # + ou -
            offset = int(m[3])
            change = m[4] if m[4] else ""
            
            if direction == "+":
                result["intron_location"] = f"intron after exon pos {exon_pos}, +{offset}bp"
            else:
                result["intron_location"] = f"intron before exon pos {exon_pos}, -{offset}bp"
            result["hgvs_c"] = f"c.{exon_pos}{direction}{offset}{change}"
            result["variant_region"] = "intronic"
            
            # Évaluer l'impact sur le splicing (seuils configurables via _splice_config)
            canonical_thresh = cls._splice_config.get("canonical_threshold", 2)
            near_thresh = cls._splice_config.get("near_threshold", 10)
            
            if offset <= canonical_thresh:
                result["splice_impact"] = "canonical splice site (high impact)"
            elif offset <= near_thresh:
                result["splice_impact"] = "near splice site (moderate impact)"
            else:
                result["splice_impact"] = "deep intronic (low impact)"
        
        # Chromosome position
        if m := cls.RE_CHR_POS.search(text):
            result["chromosome"] = m[1]
            result["chr_position"] = int(m[2])
        
        return result

    @classmethod
    def detect_variant_type(cls, text: str) -> str:
        """Détecte le type de variant."""
        if not text: return "unknown"
        text_lower = text.lower()
        
        # Intronique
        if re.search(r"c\.\d+[+-]\d+", text) or "intron" in text_lower:
            return "intronic"
        # Splice
        if "splice" in text_lower:
            return "splice"
        # Synonyme
        if "synonymous" in text_lower or "p.=" in text or "silent" in text_lower:
            return "synonymous"
        # Frameshift
        if "frameshift" in text_lower or "fs" in text_lower:
            return "frameshift"
        # Nonsense
        if "nonsense" in text_lower or "p.*" in text or "Ter" in text:
            return "nonsense"
        # Missense
        if re.search(r"p\.[A-Z][a-z]{2}\d+[A-Z][a-z]{2}", text):
            return "missense"
        # Deletion/Insertion
        if "del" in text_lower:
            return "deletion"
        if "ins" in text_lower or "dup" in text_lower:
            return "insertion"
        
        return "variant"

    @classmethod
    def impact_from_significance(cls, sig: str) -> Optional[str]:
        if not sig: return None
        s = sig.lower()
        if "pathogenic" in s and "likely" not in s: return "high"
        if "likely pathogenic" in s: return "high"
        if "benign" in s and "likely" not in s: return "low"
        if "likely benign" in s: return "low"
        if "uncertain" in s or "conflicting" in s: return "medium"
        return None

# =============================================================================
# 3. MODÈLE DE DONNÉES (Nettoyé)
# =============================================================================

@dataclass
class FieldSource:
    """Tracking de la provenance de chaque champ."""
    value: Any
    sources: List[str] = field(default_factory=list)
    
    def add_source(self, source: str):
        if source not in self.sources:
            self.sources.append(source)
    
    @property
    def count(self) -> int:
        return len(self.sources)
    
    def __repr__(self):
        return f"{self.value} [{','.join(self.sources)}]"


@dataclass
class MutationRecord:
    uniprot_id: str = ""
    position_aa: Optional[int] = None
    ref_residue: Optional[str] = None
    alt_residue: Optional[str] = None
    modification_type: str = "variant"
    impact_category: Optional[str] = None
    source_databases: List[str] = field(default_factory=list)
    validation_status: Optional[str] = None
    clinvar_ids: List[str] = field(default_factory=list)
    pubmed_refs: List[str] = field(default_factory=list)
    description: Optional[str] = None
    raw_sources: Dict[str, Any] = field(default_factory=dict)
    
    # Champs génomiques
    rsid: Optional[str] = None
    hgvs_c: Optional[str] = None  # c.1234A>G
    variant_region: Optional[str] = None  # intronic, exonic, splice, UTR
    intron_location: Optional[str] = None  # "intron 5, +15bp from exon"
    splice_impact: Optional[str] = None  # canonical, near, deep
    chromosome: Optional[str] = None
    chr_position: Optional[int] = None
    
    # Tracking multi-sources par champ
    field_sources: Dict[str, FieldSource] = field(default_factory=dict)
    
    def __post_init__(self):
        """Initialise le tracking des sources pour les champs définis à la création."""
        source = self.source_databases[0] if self.source_databases else None
        if source:
            self.init_tracking(source)
    
    def init_tracking(self, source: str):
        """Initialise le tracking des champs avec leur source d'origine."""
        trackable = {
            "position_aa": self.position_aa,
            "ref_residue": self.ref_residue,
            "alt_residue": self.alt_residue,
            "rsid": self.rsid,
            "hgvs_c": self.hgvs_c,
            "variant_region": self.variant_region,
            "intron_location": self.intron_location,
            "splice_impact": self.splice_impact,
            "chromosome": self.chromosome,
            "chr_position": self.chr_position,
            "validation_status": self.validation_status,
            "impact_category": self.impact_category,
        }
        for fname, fvalue in trackable.items():
            if fvalue is not None:
                self.track_field(fname, fvalue, source)
    
    def track_field(self, field_name: str, value: Any, source: str):
        """Track quelle source a fourni quelle valeur pour chaque champ."""
        if value is None:
            return
        if field_name not in self.field_sources:
            self.field_sources[field_name] = FieldSource(value=value, sources=[source])
        else:
            existing = self.field_sources[field_name]
            if existing.value == value:
                existing.add_source(source)
            # Si valeur différente, on garde la première mais on note le conflit
            elif f"{field_name}_alt" not in self.field_sources:
                self.field_sources[f"{field_name}_alt"] = FieldSource(value=value, sources=[source])

    @property
    def key(self) -> str:
        # Clé primaire: position AA si disponible
        if self.position_aa: 
            return f"{self.position_aa}_{self.ref_residue or '?'}_{self.alt_residue or '?'}"
        # Clé secondaire: rsID si disponible
        if self.rsid:
            return f"rs_{self.rsid}"
        # Clé tertiaire: HGVS coding (pour variants introniques)
        if self.hgvs_c:
            return f"hgvs_{self.hgvs_c}"
        # Clé quaternaire: position génomique
        if self.chr_position:
            return f"chr{self.chromosome or '?'}_{self.chr_position}"
        return f"unk_{id(self)}"
    
    @property
    def correlation_keys(self) -> List[str]:
        """Retourne toutes les clés possibles pour la corrélation multi-sources.
        
        Priorité des clés (de la plus spécifique à la plus générale):
        1. rsid - identifiant unique dbSNP
        2. mutation complète (pos+ref+alt) - très spécifique
        3. hgvs_c - nomenclature cDNA
        4. chr:position - position génomique
        5. position seule - pour les cas où ref/alt diffèrent entre sources
        """
        keys = []
        # rsID en premier (le plus fiable pour la corrélation)
        if self.rsid:
            keys.append(f"rsid:{self.rsid}")
        # Mutation complète très spécifique
        if self.position_aa and self.ref_residue and self.alt_residue:
            keys.append(f"mut:{self.position_aa}{self.ref_residue}>{self.alt_residue}")
        # hgvs_c normalisé
        if self.hgvs_c:
            # Normaliser pour la comparaison
            hgvs_norm = self.hgvs_c.split(":")[-1] if ":" in self.hgvs_c else self.hgvs_c
            keys.append(f"hgvs:{hgvs_norm}")
        # Position chromosomique
        if self.chr_position and self.chromosome:
            keys.append(f"chr:{self.chromosome}:{self.chr_position}")
        # Position AA seule (fallback pour corrélation plus souple)
        if self.position_aa:
            keys.append(f"pos:{self.position_aa}")
        return keys

    def merge(self, other: 'MutationRecord') -> None:
        """Fusionne deux records en trackant les sources de chaque champ."""
        other_source = other.source_databases[0] if other.source_databases else "unknown"
        
        # Champs protéiques avec tracking
        if other.position_aa:
            self.track_field("position_aa", other.position_aa, other_source)
            if not self.position_aa: self.position_aa = other.position_aa
        if other.ref_residue:
            self.track_field("ref_residue", other.ref_residue, other_source)
            if not self.ref_residue: self.ref_residue = other.ref_residue
        if other.alt_residue:
            self.track_field("alt_residue", other.alt_residue, other_source)
            if not self.alt_residue: self.alt_residue = other.alt_residue
        
        # Champs génomiques avec tracking
        if other.rsid:
            self.track_field("rsid", other.rsid, other_source)
            if not self.rsid: self.rsid = other.rsid
        if other.hgvs_c:
            self.track_field("hgvs_c", other.hgvs_c, other_source)
            if not self.hgvs_c: self.hgvs_c = other.hgvs_c
        if other.variant_region:
            self.track_field("variant_region", other.variant_region, other_source)
            if not self.variant_region: self.variant_region = other.variant_region
        if other.intron_location:
            self.track_field("intron_location", other.intron_location, other_source)
            if not self.intron_location: self.intron_location = other.intron_location
        if other.splice_impact:
            self.track_field("splice_impact", other.splice_impact, other_source)
            if not self.splice_impact: self.splice_impact = other.splice_impact
        if other.chromosome:
            self.track_field("chromosome", other.chromosome, other_source)
            if not self.chromosome: self.chromosome = other.chromosome
        if other.chr_position:
            self.track_field("chr_position", other.chr_position, other_source)
            if not self.chr_position: self.chr_position = other.chr_position
        
        # Validation et description avec tracking
        if other.validation_status:
            self.track_field("validation_status", other.validation_status, other_source)
            if not self.validation_status: self.validation_status = other.validation_status
        if other.description:
            self.track_field("description", other.description, other_source)
            if not self.description: self.description = other.description
        
        # Listes (union)
        self.source_databases = sorted(list(set(self.source_databases + other.source_databases)))
        self.clinvar_ids = sorted(list(set(self.clinvar_ids + other.clinvar_ids)))
        self.pubmed_refs = sorted(list(set(self.pubmed_refs + other.pubmed_refs)))
        
        # Impact (priorité au plus élevé) avec tracking
        priority = {"high": 3, "medium": 2, "low": 1}
        if other.impact_category:
            self.track_field("impact_category", other.impact_category, other_source)
            if priority.get(other.impact_category, 0) > priority.get(self.impact_category, 0):
                self.impact_category = other.impact_category
        
        # Fusionner les field_sources de l'autre record
        for field, fs in other.field_sources.items():
            if field not in self.field_sources:
                self.field_sources[field] = fs
            else:
                for src in fs.sources:
                    self.field_sources[field].add_source(src)
        
        self.raw_sources.update(other.raw_sources)

    def to_dict(self) -> Dict[str, Any]:
        mutation = ""
        if self.ref_residue or self.alt_residue: 
            ref = self.ref_residue or "?"
            alt = self.alt_residue or "?"
            mutation = f"{ref}→{alt}"
        
        # Construire la localisation détaillée
        location_detail = ""
        if self.variant_region == "intronic" and self.intron_location:
            location_detail = f"{self.intron_location}"
            if self.splice_impact:
                location_detail += f" [{self.splice_impact}]"
        elif self.hgvs_c:
            # Normaliser hgvs_c pour l'affichage
            location_detail = VariantParser.normalize_hgvs_c(self.hgvs_c)
        
        # Normaliser hgvs_c pour l'export
        hgvs_c_normalized = VariantParser.normalize_hgvs_c(self.hgvs_c) if self.hgvs_c else ""
        
        # Compter les confirmations multi-sources pour les champs clés
        key_fields = ["position_aa", "rsid", "ref_residue", "alt_residue", "validation_status", "hgvs_c", "variant_region"]
        confirmations = sum(1 for f in key_fields if f in self.field_sources and self.field_sources[f].count > 1)
        
        # Construire le détail des sources par champ (format compact pour Excel)
        source_detail = {}
        source_summary = []  # Format lisible: "pos:UniProt+ClinVar, rsid:ClinVar"
        for fname, fs in self.field_sources.items():
            if fs.count > 1 and not fname.endswith("_alt"):
                source_detail[fname] = {"value": fs.value, "sources": fs.sources, "count": fs.count}
                source_summary.append(f"{fname}:{'+'.join(fs.sources)}")
        
        # Format JSON pour field_sources (sérialisable)
        field_sources_json = json.dumps(source_detail, ensure_ascii=False) if source_detail else ""
        
        return {
            "uniprot_id": self.uniprot_id, 
            "position": self.position_aa, 
            "mutation": mutation,
            "rsid": self.rsid or "",
            "hgvs_c": hgvs_c_normalized,
            "type": self.modification_type, 
            "region": self.variant_region or ("exonic" if self.position_aa else ""),
            "location_detail": location_detail,
            "chromosome": self.chromosome or "",
            "chr_position": self.chr_position or "",
            "impact": self.impact_category,
            "sources": ",".join(self.source_databases),
            "source_count": len(self.source_databases),
            "confirmations": confirmations,  # Nb de champs confirmés par 2+ sources
            "confirmed_fields": "; ".join(source_summary) if source_summary else "",  # Format lisible
            "validation": self.validation_status,
            "description": self.description,
            "pubmed_ids": ",".join(set(self.pubmed_refs)) or "",
            "clinvar_id": ",".join(set(self.clinvar_ids)) or "",
            "field_sources": field_sources_json,  # JSON string pour compatibilité Excel
        }

# =============================================================================
# 4. INFRASTRUCTURE
# =============================================================================

class Cache:
    def __init__(self, path: str, timeout: int = 30):
        self.path, self._lock = path, threading.RLock()
        self.timeout = timeout
        with self._conn() as c:
            c.execute("PRAGMA journal_mode=WAL")
            c.execute("CREATE TABLE IF NOT EXISTS cache (key TEXT PRIMARY KEY, url TEXT, params TEXT, response TEXT, timestamp REAL)")

    @contextmanager
    def _conn(self):
        conn = sqlite3.connect(self.path, check_same_thread=False, timeout=self.timeout)
        try: yield conn
        finally: conn.close()

    def _key(self, url: str, params: Dict = None) -> str:
        return hashlib.sha256(json.dumps({"url": url, "params": params or {}}, sort_keys=True).encode()).hexdigest()

    def get(self, url: str, params: Dict = None) -> Any:
        with self._lock, self._conn() as c:
            row = c.execute("SELECT response FROM cache WHERE key=?", (self._key(url, params),)).fetchone()
            if row:
                try: return json.loads(row[0])
                except json.JSONDecodeError: return row[0]
        return None

    def set(self, url: str, params: Dict, data: Any) -> None:
        with self._lock, self._conn() as c:
            c.execute("REPLACE INTO cache (key,url,params,response,timestamp) VALUES (?,?,?,?,?)",
                      (self._key(url, params), url, json.dumps(params, default=str), json.dumps(data, default=str), time.time()))
            c.commit()

class HTTPClient:
    def __init__(self, config: Dict):
        self.session = requests.Session()
        self.config = config
        http_cfg = config.get("http_client", {})
        # Codes HTTP pour retry (configurable)
        status_codes = http_cfg.get("retry_status_codes", [429, 500, 502, 503, 504, 404])
        retry = Retry(total=http_cfg.get("retries", 3), backoff_factor=http_cfg.get("backoff_factor", 0.5),
                      status_forcelist=status_codes, allowed_methods=["GET"])
        self.session.mount("https://", HTTPAdapter(max_retries=retry))
        self.limits = http_cfg.get("rate_limits", {})
        self._locks, self._last = {}, {}

    def get(self, url: str, params: Dict = None) -> Any:
        host = requests.utils.urlparse(url).netloc
        if limit := self.limits.get(host):
            lock = self._locks.setdefault(host, threading.Lock())
            with lock:
                wait = 1.0 / limit - (time.time() - self._last.get(host, 0))
                if wait > 0: time.sleep(wait)
                self._last[host] = time.time()
        
        http_cfg = self.config.get("http_client", {})
        resp = self.session.get(url, params=params, timeout=(http_cfg.get("connect_timeout", 10), http_cfg.get("read_timeout", 30)))
        resp.raise_for_status()
        try: return resp.json()
        except ValueError: return resp.text

# =============================================================================
# 5. ADAPTERS (Seulement les sources pertinentes)
# =============================================================================

class BaseAdapter(ABC):
    def __init__(self, client: HTTPClient, cache: Cache, config: Dict, uid_context: str = ""):
        # Client HTTP, cache et config partagés
        self.client, self.cache, self.config = client, cache, config
        self.log = logger.getChild(self.__class__.__name__)
        self.uid_context = uid_context
        # Pré-calculs depuis la config pour éviter les appels répétés
        self.fetch_all = bool(self.config.get("sources", {}).get("fetch_all", False))
        self.limits = self.config.get("limits", {})
        self.batch_size = self.limits.get("batch_size", 20)

    def _log(self, level, msg, *args, **kwargs):
        self.log.log(level, msg, *args, extra={'uid': self.uid_context}, **kwargs)

    def _fetch(self, url: str, params: Dict = None) -> Any:
        """Fetch + cache wrapper avec logging centralisé."""
        cached = self.cache.get(url, params)
        if cached is not None: return cached
        try:
            data = self.client.get(url, params)
            self.cache.set(url, params, data)
            return data
        except Exception as e:
            self._log(logging.WARNING, "Fetch failed %s: %s", url, e)
            return None

    # Helper utilitaires
    def get_limit(self, name: str, default: int) -> Optional[int]:
        """Retourne la limite configurée ou None si fetch_all est activé."""
        if self.fetch_all:
            return None
        return self.limits.get(name, default)

    def slice_items(self, items: List[Any], limit_name: str, default: int) -> List[Any]:
        """Retourne la liste tronquée si nécessaire en fonction des limites."""
        if items is None:
            return []
        limit = self.get_limit(limit_name, default)
        return list(items) if limit is None else list(items)[:limit]

    def make_record(self, source: str, **kwargs) -> MutationRecord:
        """Constructeur simplifié de MutationRecord (centralise source et raw_sources)."""
        sources = kwargs.pop('source_databases', [source])
        raw = kwargs.pop('raw_sources', {})
        rec = MutationRecord(source_databases=sources, raw_sources=raw, **kwargs)
        rec.init_tracking(source)
        return rec

    @abstractmethod
    def fetch_data(self, uniprot_id: str, gene_symbol: str = None) -> List[MutationRecord]: pass


class UniProtAdapter(BaseAdapter):
    def fetch_data(self, uniprot_id: str, gene_symbol: str = None) -> List[MutationRecord]:
        data = self._fetch(f"https://rest.uniprot.org/uniprotkb/{uniprot_id}.json")
        if not data: return []
        records = []
        for feat in data.get("features", []):
            feat_type = (feat.get("type") or "").lower()
            if feat_type not in ("variant", "mutagenesis", "natural variant", "sequence variant"): continue
            pos, ref, alt = None, None, None
            loc = feat.get("location", {})
            start = loc.get("start", {})
            pos = start.get("value") if isinstance(start, dict) else start
            
            alt_seq = feat.get("alternativeSequence", {})
            ref = alt_seq.get("originalSequence") or feat.get("original")
            alt = (alt_seq.get("alternativeSequences") or [None])[0] or feat.get("variation")
            
            description = feat.get("description", "")
            if pos is None: 
                pos, ref, alt = VariantParser.parse_protein(description)
            
            # Extraire rsID de la description
            rsid = VariantParser.parse_rsid(description)
            
            pmids = [str(e["id"]) for e in feat.get("evidences", []) if e.get("source") == "PubMed" and e.get("id")]
            records.append(self.make_record(
                "UniProt",
                uniprot_id=uniprot_id, position_aa=pos, ref_residue=ref, alt_residue=alt,
                modification_type=feat.get("type", "variant"), 
                impact_category="high" if pmids else None,
                validation_status="experimental" if pmids else None,
                pubmed_refs=pmids, 
                description=description, 
                rsid=rsid,
                variant_region="exonic" if pos else None,
                raw_sources={"uniprot": feat}
            ))
        self._log(logging.INFO, "UniProt features fetched: %d", len(records))
        return records

    def get_gene_info(self, uniprot_id: str) -> Dict[str, Optional[str]]:
        """Retourne les identifiants du gène: symbol, ncbi_gene_id."""
        data = self._fetch(f"https://rest.uniprot.org/uniprotkb/{uniprot_id}.json")
        if not data: return {"symbol": None, "ncbi_gene_id": None}
        
        result = {"symbol": None, "ncbi_gene_id": None}
        
        # Gene Symbol
        try: 
            result["symbol"] = data.get("genes", [{}])[0].get("geneName", {}).get("value")
        except: pass
        
        # NCBI Gene ID depuis les cross-references
        for xref in data.get("uniProtKBCrossReferences", []):
            if xref.get("database") == "GeneID":
                result["ncbi_gene_id"] = xref.get("id")
                break
        
        return result


class ClinVarAdapter(BaseAdapter):
    BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
    
    def fetch_data(self, uniprot_id: str, gene_id: str = None, known_rsids: List[str] = None) -> List[MutationRecord]:
        """Recherche ClinVar par Gene ID (NCBI) et rsIDs connus.
        
        Args:
            uniprot_id: UniProt accession
            gene_id: NCBI Gene ID (ex: 10599 pour SLCO1B1)
            known_rsids: Liste de rsIDs connus d'UniProt
        """
        uids = set()
        
        # fetch_all global: récupère tous les variants (ignore les limites)
        search_limit = 100000 if self.fetch_all else self.get_limit("clinvar_search_limit", 200)
        
        # 1. Recherche par rsIDs connus en batch (prioritaire - meilleure corrélation)
        if known_rsids:
            for i in range(0, len(known_rsids), self.batch_size):
                batch = known_rsids[i:i+self.batch_size]
                rs_terms = " OR ".join(f"rs{rsid.replace('rs', '')}[All Fields]" for rsid in batch if rsid)
                if rs_terms:
                    data = self._fetch(f"{self.BASE_URL}/esearch.fcgi",
                                      {"db": "clinvar", "term": rs_terms, "retmode": "json", "retmax": str(search_limit)})
                    if data:
                        uids.update(data.get("esearchresult", {}).get("idlist", []))
        
        # 2. Recherche par NCBI Gene ID (identifiant stable)
        if gene_id:
            self._log(logging.INFO, "ClinVar search by Gene ID: %s", gene_id)
            data = self._fetch(f"{self.BASE_URL}/esearch.fcgi", 
                              {"db": "clinvar", "term": f"{gene_id}[geneid]", "retmode": "json", "retmax": str(search_limit)})
            if data: 
                uids.update(data.get("esearchresult", {}).get("idlist", []))
        
        if not uids: 
            return []
        
        self._log(logging.INFO, "ClinVar UIDs found: %d", len(uids))
        clinvar_limit = self.get_limit("clinvar_max_records", 50)
        uid_list = list(uids) if clinvar_limit is None else list(uids)[:clinvar_limit]
        
        # Utiliser esummary en batch (paralléliser les requêtes par petits lots)
        records = []
        batch_calls = []
        for i in range(0, len(uid_list), self.batch_size):
            batch = uid_list[i:i+self.batch_size]
            batch_calls.append((batch, {"db": "clinvar", "id": ",".join(batch), "retmode": "json"}))

        if batch_calls:
            inner_workers = min(2, self.limits.get('max_workers', 4))
            from concurrent.futures import ThreadPoolExecutor, as_completed
            with ThreadPoolExecutor(max_workers=inner_workers) as rex:
                fut_map = {rex.submit(self._fetch, f"{self.BASE_URL}/esummary.fcgi", params): batch for (batch, params) in batch_calls}
                for fut in as_completed(fut_map):
                    batch = fut_map[fut]
                    try:
                        summary = fut.result()
                    except Exception:
                        self._log(logging.WARNING, "ClinVar: esummary fetch failed for batch")
                        continue
                    if summary and "result" in summary:
                        for uid in batch:
                            if uid in summary["result"]:
                                rec = self._parse_summary(summary["result"][uid], uid, uniprot_id)
                                if rec:
                                    records.append(rec)

        self._log(logging.INFO, "ClinVar records parsed: %d", len(records))
        return records

    def _parse_summary(self, data: Dict, clinvar_id: str, uniprot_id: str) -> Optional[MutationRecord]:
        """Parse un résultat esummary ClinVar."""
        if not data or data.get("error"):
            return None
        
        # === Classification ===
        germline = data.get("germline_classification", {})
        sig = germline.get("description", "")
        
        # === rsID et Position ===
        rsid = None
        pos, ref, alt = None, None, None
        hgvs_c = None
        hgvs_p = None  # p.Gly611Ter, p.Asn130Asp, etc.
        
        for var_set in data.get("variation_set", []):
            # dbSNP
            for xref in var_set.get("variation_xrefs", []):
                if xref.get("db_source") == "dbSNP":
                    rsid = f"rs{xref.get('db_id', '')}"
                    break
            
            # Protein change depuis aliases (M235T, L432V, etc.)
            for alias in var_set.get("aliases", []):
                if not pos:
                    pos, ref, alt = VariantParser.parse_protein(alias)
                # HGVS coding
                if ":c." in alias and not hgvs_c:
                    hgvs_c = VariantParser.normalize_hgvs_c(alias)
                # HGVS protein (p.Xxx###Xxx ou p.Xxx###Ter/*)
                if "p." in alias and not hgvs_p:
                    hgvs_p = alias
            
            # cDNA change direct
            if not hgvs_c and var_set.get("cdna_change"):
                hgvs_c = VariantParser.normalize_hgvs_c(var_set.get("cdna_change"))
            
            # Variation name comme fallback
            var_name = var_set.get("variation_name", "")
            if not pos:
                pos, ref, alt = VariantParser.parse_protein(var_name)
            # Extraire hgvs_c du variation_name si pas encore trouvé
            if not hgvs_c and ":c." in var_name:
                hgvs_c = VariantParser.normalize_hgvs_c(var_name)
            # Extraire hgvs_p du variation_name (p.Xxx###Xxx)
            if not hgvs_p and "(p." in var_name:
                if m := re.search(r"\(p\.[^)]+\)", var_name):
                    hgvs_p = m.group(0).strip("()")
            
            # Position chromosomique
            chr_pos = None
            chromosome = None
            for loc in var_set.get("variation_loc", []):
                if loc.get("assembly_name") == "GRCh38":
                    chromosome = loc.get("chr")
                    try:
                        chr_pos = int(loc.get("start", 0))
                    except (ValueError, TypeError):
                        pass
                    break
        
        # Extraire hgvs_c depuis le titre si pas encore trouvé
        title = data.get("title", "")
        if not hgvs_c and ":c." in title:
            hgvs_c = VariantParser.normalize_hgvs_c(title)
        # Extraire hgvs_p depuis le titre
        if not hgvs_p and "(p." in title:
            if m := re.search(r"\(p\.[^)]+\)", title):
                hgvs_p = m.group(0).strip("()")
        
        # Si on a hgvs_p mais pas alt, parser le alt depuis hgvs_p
        if hgvs_p and pos and ref and not alt:
            # Format: p.Gly611Ter -> alt = * (stop)
            # Format: p.Gly611= -> synonyme
            if m := re.search(r"p\.[A-Z][a-z]{2}(\d+)([A-Z][a-z]{2}|Ter|=|\*)", hgvs_p, re.I):
                alt_raw = m.group(2)
                if alt_raw.lower() == "ter" or alt_raw == "*":
                    alt = "*"
                elif alt_raw == "=":
                    alt = ref  # Synonyme
                else:
                    alt = VariantParser.AA_3TO1.get(alt_raw.capitalize(), alt_raw)
        
        # Parsing génomique/intronique (utiliser hgvs_c ou le title)
        genomic_info = VariantParser.parse_genomic_location(hgvs_c or title)
        variant_region = genomic_info.get("variant_region")
        intron_location = genomic_info.get("intron_location")
        splice_impact = genomic_info.get("splice_impact")
        
        # Si pas de region détectée mais hgvs_c présent, c'est exonique
        if not variant_region and hgvs_c and not re.search(r"c\.\d+[+-]\d+", hgvs_c):
            variant_region = "exonic"
        
        # Variant type
        variant_type = VariantParser.detect_variant_type(title + " " + (hgvs_c or ""))
        
        # Accession
        accession = data.get("accession", clinvar_id)
        
        # Description enrichie
        desc_parts = [sig] if sig else []
        if variant_region == "intronic" and intron_location:
            desc_parts.append(f"[{intron_location}]")
        if splice_impact and "canonical" in splice_impact.lower():
            desc_parts.append(f"⚠️ {splice_impact}")
        
        return self.make_record(
            "ClinVar",
            uniprot_id=uniprot_id,
            position_aa=pos,
            ref_residue=ref,
            alt_residue=alt,
            modification_type=variant_type if variant_type != "variant" else "clinical",
            impact_category=VariantParser.impact_from_significance(sig),
            validation_status=sig,
            clinvar_ids=[accession],
            pubmed_refs=[],  # esummary ne contient pas les PMIDs directement
            description=" | ".join(desc_parts) if desc_parts else None,
            rsid=rsid,
            hgvs_c=hgvs_c,
            variant_region=variant_region,
            intron_location=intron_location,
            splice_impact=splice_impact,
            chromosome=chromosome,
            chr_position=chr_pos,
            raw_sources={"clinvar_summary": data.get("title", "")}
        )


class PharmGKBAdapter(BaseAdapter):
    """Adapter PharmGKB avec les endpoints corrects de l'API REST.
    
    Endpoints utilisés:
    - /data/gene?crossReferences.resource=UniProtKB&crossReferences.resourceId=X → Info gène par UniProt ID
    - /report/connectedObjects/{gene_id}/Variant → Variants connectés
    - /data/variant/?symbol=rsXXX → Détails variant par rsID
    - /data/clinicalAnnotation?location.genes.accessionId=X → Annotations cliniques
    - /data/variantAnnotation?location.genes.accessionId=X → Annotations de variants
    """
    BASE_URL = "https://api.pharmgkb.org/v1"
    
    # Valeurs par défaut (surchargeables via config)
    DEFAULT_LEVEL_SCORES = {"1A": 5, "1B": 4, "2A": 3, "2B": 2, "3": 1, "4": 0, "N/A": -1}
    
    @property
    def LEVEL_RANK(self) -> Dict[str, int]:
        """Retourne les scores de niveau depuis la config ou valeurs par défaut."""
        return self.config.get("pharmgkb", {}).get("level_scores", self.DEFAULT_LEVEL_SCORES)

    def _extract_data(self, response: Any) -> Any:
        """Extrait les données de la réponse (format {"data": [...]} ou directement [...])."""
        if isinstance(response, dict):
            return response.get('data', response)
        return response

    def fetch_data(self, uniprot_id: str, gene_id: str = None) -> List[MutationRecord]:
        """Recherche PharmGKB par UniProt ID (cross-reference).
        
        Args:
            uniprot_id: UniProt accession (ex: Q9Y6L6)
            gene_id: Non utilisé, gardé pour compatibilité signature
        """
        records = []
        
        # 1. Recherche du gène par UniProt ID (cross-reference)
        gene_data = self._fetch(f"{self.BASE_URL}/data/gene", 
                               {"crossReferences.resource": "UniProtKB", 
                                "crossReferences.resourceId": uniprot_id})
        if not gene_data: 
            self._log(logging.DEBUG, "PharmGKB: No gene found for UniProt %s", uniprot_id)
            return []
        gene_list = self._extract_data(gene_data)
        if not gene_list or not isinstance(gene_list, list): 
            return []
        gene = gene_list[0]
        
        pharmgkb_gene_id = gene.get("id")
        gene_symbol = gene.get("symbol")
        self._log(logging.INFO, "PharmGKB Gene ID: %s (symbol: %s)", pharmgkb_gene_id, gene_symbol)
        
        # Limites configurables (fetch_all = pas de limite)
        limits = self.limits
        limit_connected = self.get_limit("pharmgkb_connected_variants", 20)
        limit_clinical = self.get_limit("pharmgkb_clinical_annotations", 30)
        limit_variant = self.get_limit("pharmgkb_variant_annotations", 50)
        
        # 2. Variants Connectés au gène (via /report/connectedObjects)
        connected = self._fetch(f"{self.BASE_URL}/report/connectedObjects/{pharmgkb_gene_id}/Variant")
        if connected:
            connected_list = self._extract_data(connected) if isinstance(connected, dict) else connected
            self._log(logging.INFO, "PharmGKB Connected Variants: %d", len(connected_list) if connected_list else 0)
            
            # Pour chaque variant connecté, récupérer les détails (paralléliser les fetchs pour accélérer)
            items = self.slice_items(connected_list, 'pharmgkb_connected_variants', 20)
            rs_tasks = []
            for conn in items:
                var_obj = conn.get('connectedObject', {})
                rsid = var_obj.get('symbol', '')
                conn_types = conn.get('connectionTypes', [])
                if rsid and rsid.startswith('rs'):
                    rs_tasks.append((rsid, conn_types))

            if rs_tasks:
                # Limiter le parallélisme interne pour respecter les rate limits
                inner_workers = min(4, self.limits.get('max_workers', 4))
                from concurrent.futures import ThreadPoolExecutor, as_completed
                with ThreadPoolExecutor(max_workers=inner_workers) as rex:
                    f_map = {rex.submit(self._fetch, f"{self.BASE_URL}/data/variant/", {"symbol": rsid, "view": "max"}): (rsid, ct) for rsid, ct in rs_tasks}
                    for fut in as_completed(f_map):
                        rsid, conn_types = f_map[fut]
                        try:
                            var_details = fut.result()
                        except Exception:
                            self._log(logging.WARNING, "PharmGKB variant fetch failed for %s", rsid)
                            continue
                        if var_details:
                            var_list = self._extract_data(var_details)
                            if var_list and isinstance(var_list, list):
                                var = var_list[0]
                                rec = self._parse_variant(var, uniprot_id, conn_types)
                                if rec:
                                    records.append(rec)
        
        # 3. Annotations Cliniques par Gene ID PharmGKB (niveau de preuve par médicament)
        clinical_anns = self._fetch(f"{self.BASE_URL}/data/clinicalAnnotation", 
                                   {"location.genes.accessionId": pharmgkb_gene_id, "view": "max"})
        if clinical_anns:
            ann_list = self._extract_data(clinical_anns)
            self._log(logging.INFO, "PharmGKB Clinical Annotations: %d", len(ann_list) if ann_list else 0)
            
            # Parser les annotations pour extraire les variants spécifiques
            clinical_items = self.slice_items(ann_list, 'pharmgkb_clinical_annotations', 30)
            if clinical_items:
                inner_workers = min(4, self.limits.get('max_workers', 4))
                from concurrent.futures import ThreadPoolExecutor, as_completed
                with ThreadPoolExecutor(max_workers=inner_workers) as rex:
                    futs = {rex.submit(self._parse_clinical_annotation, ann, uniprot_id): ann for ann in clinical_items}
                    for fut in as_completed(futs):
                        try:
                            rec = fut.result()
                            if rec:
                                records.append(rec)
                        except Exception:
                            self._log(logging.WARNING, "PharmGKB: parsing clinical annotation failed")
            
        
        # 4. Annotations de Variants par Gene ID PharmGKB (plus détaillées)
        var_anns = self._fetch(f"{self.BASE_URL}/data/variantAnnotation", 
                              {"location.genes.accessionId": pharmgkb_gene_id, "view": "base"})
        if var_anns:
            vann_list = self._extract_data(var_anns)
            self._log(logging.INFO, "PharmGKB Variant Annotations: %d", len(vann_list) if vann_list else 0)
            
            # Parser les annotations avec positions protéiques
            variant_items = self.slice_items(vann_list, 'pharmgkb_variant_annotations', 50)
            if variant_items:
                inner_workers = min(4, self.limits.get('max_workers', 4))
                from concurrent.futures import ThreadPoolExecutor, as_completed
                with ThreadPoolExecutor(max_workers=inner_workers) as rex:
                    futs = {rex.submit(self._parse_variant_annotation, vann, uniprot_id): vann for vann in variant_items}
                    for fut in as_completed(futs):
                        try:
                            rec = fut.result()
                            if rec:
                                records.append(rec)
                        except Exception:
                            self._log(logging.WARNING, "PharmGKB: parsing variant annotation failed")
        
        # 5. Record Global d'enrichissement (niveau de preuve du gène)
        best_level = self._get_best_evidence_level(ann_list if clinical_anns else [])
        guideline_names = self._get_guideline_names(pharmgkb_gene_id)
        
        global_desc = f"PharmGKB Gene Lvl {best_level}. Guidelines: {guideline_names or 'None'}"
        global_record = self.make_record(
            "PharmGKB",
            uniprot_id=uniprot_id,
            position_aa=None,
            description=global_desc,
            impact_category=self._map_level_to_impact(best_level)
        )
        records.append(global_record)
        
        self._log(logging.INFO, "PharmGKB records total: %d", len(records))
        return records

    def _parse_variant(self, var: Dict, uniprot_id: str, connection_types: List[str]) -> Optional[MutationRecord]:
        """Parse un variant depuis /data/variant/."""
        rsid = var.get('symbol', '')
        if not rsid.startswith('rs'):
            return None
        
        # Extraire la position protéique depuis locations
        pos, ref, alt = None, None, None
        for loc in var.get('locations', []):
            if loc.get('proteinChange'):
                pos, ref, alt = VariantParser.parse_protein(loc['proteinChange'])
                if pos:
                    break
        
        # Classification et signification clinique
        change_class = var.get('changeClassification', '')
        clinical_sig = var.get('clinicalSignificance', '')
        
        # Niveau d'impact basé sur les types de connexion
        impact = "medium"
        if "Guideline Annotation" in connection_types or "Label Annotation" in connection_types:
            impact = "high"
        
        desc = f"PharmGKB: {change_class}"
        if clinical_sig:
            desc += f" ({clinical_sig})"
        if connection_types:
            desc += f" [{', '.join(connection_types)}]"
        
        return self.make_record(
            "PharmGKB",
            uniprot_id=uniprot_id,
            position_aa=pos,
            ref_residue=ref,
            alt_residue=alt,
            rsid=rsid,
            modification_type=change_class.lower() if change_class else "variant",
            description=desc,
            impact_category=impact,
            raw_sources={"pharmgkb_variant": var.get('id')}
        )

    def _parse_clinical_annotation(self, ann: Dict, uniprot_id: str) -> Optional[MutationRecord]:
        """Parse une annotation clinique pour extraire variant et niveau de preuve."""
        # Niveau de preuve
        level_obj = ann.get('levelOfEvidence', {})
        level = level_obj.get('term') if isinstance(level_obj, dict) else str(level_obj)
        
        # Location et variants
        location = ann.get('location', {})
        variants = location.get('variants', []) if isinstance(location, dict) else []
        
        if not variants:
            return None
        
        # Premier variant avec rsID
        rsid = None
        pos, ref, alt = None, None, None
        for v in variants:
            v_symbol = v.get('symbol', '')
            if v_symbol.startswith('rs'):
                rsid = v_symbol
                # Essayer de parser la position protéique
                if v.get('proteinChange'):
                    pos, ref, alt = VariantParser.parse_protein(v['proteinChange'])
                break
        
        if not rsid:
            return None
        
        # Médicaments associés
        chemicals = [c.get('name') for c in ann.get('relatedChemicals', [])]
        
        desc = f"PharmGKB Lvl {level}"
        if chemicals:
            desc += f": {', '.join(chemicals[:3])}"
        
        return MutationRecord(
            uniprot_id=uniprot_id,
            position_aa=pos,
            ref_residue=ref,
            alt_residue=alt,
            rsid=rsid,
            source_databases=["PharmGKB"],
            description=desc,
            impact_category=self._map_level_to_impact(level),
            validation_status=f"Level {level}",
            raw_sources={"pharmgkb_clinical_ann": ann.get('id')}
        )

    def _parse_variant_annotation(self, vann: Dict, uniprot_id: str) -> Optional[MutationRecord]:
        """Parse une annotation de variant avec phrase descriptive."""
        location = vann.get('location', {})
        variants = location.get('variants', []) if isinstance(location, dict) else []
        
        rsid = None
        pos, ref, alt = None, None, None
        
        for v in variants:
            v_symbol = v.get('symbol', '')
            if v_symbol.startswith('rs'):
                rsid = v_symbol
                break
        
        if not rsid:
            return None
        
        # Description depuis la phrase
        sentence = vann.get('sentence', '')[:200]
        
        return MutationRecord(
            uniprot_id=uniprot_id,
            position_aa=pos,
            ref_residue=ref,
            alt_residue=alt,
            rsid=rsid,
            source_databases=["PharmGKB"],
            description=f"PharmGKB: {sentence}" if sentence else None,
            impact_category="medium",
            raw_sources={"pharmgkb_var_ann": vann.get('id')}
        )

    def _get_best_evidence_level(self, annotations: List[Dict]) -> str:
        """Trouve le meilleur niveau de preuve parmi les annotations."""
        best = "N/A"
        best_rank = -1
        for ann in (annotations or []):
            level_obj = ann.get('levelOfEvidence', {})
            level = level_obj.get('term') if isinstance(level_obj, dict) else str(level_obj)
            rank = self.LEVEL_RANK.get(level, -1)
            if rank > best_rank:
                best_rank = rank
                best = level
        return best

    def _get_guideline_names(self, gene_id: str) -> str:
        """Récupère les noms des guidelines pour le gène."""
        guides = self._fetch(f"{self.BASE_URL}/data/guidelineAnnotation", 
                            {"relatedGenes.accessionId": gene_id})
        if guides:
            guide_list = self._extract_data(guides)
            if guide_list:
                names = [g.get('name', '')[:50] for g in guide_list[:3]]
                return "; ".join(names)
        return ""

    def _map_level_to_impact(self, level: str) -> str:
        """Convertit un niveau PharmGKB en impact (seuils configurables)."""
        pharmgkb_cfg = self.config.get("pharmgkb", {})
        high_thresh = pharmgkb_cfg.get("high_threshold", 4)
        medium_thresh = pharmgkb_cfg.get("medium_threshold", 2)
        
        r = self.LEVEL_RANK.get(level, -1)
        if r >= high_thresh: return "high"
        elif r >= medium_thresh: return "medium"
        elif r >= 0: return "low"
        return "medium"

# =============================================================================
# 6. AGGREGATEUR (Clean Logic)
# =============================================================================

class CorrelationEngine:
    """Moteur de corrélation multi-sources basé sur index multi-clé.
    
    Stratégie de matching configurable:
    - rsid: identifiant unique dbSNP (très fiable)
    - mut: mutation complète (pos+ref+alt) 
    - hgvs: nomenclature cDNA
    - chr: position chromosomique
    - pos: position AA seule (fallback)
    
    Config correlation:
        min_criteria: 1  # Nombre minimum de critères communs pour fusionner
        enabled_keys: [rsid, mut, hgvs, chr, pos]  # Clés actives
        key_weights: {rsid: 5, mut: 4, ...}  # Poids des critères
    """
    
    # Critères de corrélation par défaut (surchargeables via config)
    DEFAULT_KEY_WEIGHTS = {"rsid": 5, "mut": 4, "hgvs": 3, "chr": 2, "pos": 1}
    
    def __init__(self, config: Dict = None):
        self.key_to_record: Dict[str, MutationRecord] = {}
        self.correlation_index: Dict[str, str] = {}
        self.stats = {"total": 0, "merged": 0, "new": 0, "global": 0}
        
        # Configuration
        corr_config = (config or {}).get("correlation", {})
        self.min_criteria = corr_config.get("min_criteria", 1)
        self.enabled_keys = set(corr_config.get("enabled_keys", ["rsid", "mut", "hgvs", "chr", "pos"]))
        self.KEY_WEIGHTS = corr_config.get("key_weights", self.DEFAULT_KEY_WEIGHTS)
    
    def _filter_keys(self, keys: List[str]) -> List[str]:
        """Filtre les clés selon la configuration."""
        return [k for k in keys if k.split(":")[0] in self.enabled_keys]
    
    def add_record(self, rec: MutationRecord) -> bool:
        """Ajoute un record avec corrélation automatique multi-clé."""
        if not rec.source_databases and not any([rec.position_aa, rec.rsid, rec.hgvs_c, rec.chr_position]):
            self.stats["global"] += 1
            return False
        
        self.stats["total"] += 1
        corr_keys = self._filter_keys(rec.correlation_keys)
        
        # Chercher les matches et compter les critères communs
        matched_key = None
        match_type = None
        match_count = 0
        
        for corr_key in corr_keys:
            if corr_key in self.correlation_index:
                candidate_key = self.correlation_index[corr_key]
                if candidate_key == matched_key:
                    match_count += 1
                elif matched_key is None:
                    matched_key = candidate_key
                    match_type = corr_key.split(":")[0]
                    match_count = 1
        
        # Vérifier le nombre minimum de critères
        if matched_key and match_count >= self.min_criteria and matched_key in self.key_to_record:
            existing = self.key_to_record[matched_key]
            
            # Vérification de cohérence pour match par position seule
            if match_type == "pos" and rec.ref_residue and existing.ref_residue:
                if rec.ref_residue != existing.ref_residue:
                    return self._add_new_record(rec, corr_keys)
            
            # Fusion
            existing.merge(rec)
            for corr_key in corr_keys:
                if corr_key not in self.correlation_index:
                    self.correlation_index[corr_key] = matched_key
            self.stats["merged"] += 1
            return True
        else:
            return self._add_new_record(rec, corr_keys)
    
    def _add_new_record(self, rec: MutationRecord, corr_keys: List[str]) -> bool:
        """Ajoute un nouveau record non-fusionné."""
        unique_key = rec.key
        if unique_key in self.key_to_record:
            unique_key = f"{unique_key}_{len(self.key_to_record)}"
        self.key_to_record[unique_key] = rec
        for corr_key in corr_keys:
            if corr_key not in self.correlation_index:
                self.correlation_index[corr_key] = unique_key
        self.stats["new"] += 1
        return False
    
    def get_records(self) -> List[MutationRecord]:
        return list(self.key_to_record.values())
    
    def get_correlation_stats(self) -> Dict:
        """Retourne les statistiques de corrélation."""
        multi_source = sum(1 for r in self.key_to_record.values() if len(r.source_databases) > 1)
        return {
            **self.stats,
            "unique_records": len(self.key_to_record),
            "multi_source_records": multi_source,
            "correlation_keys": len(self.correlation_index),
            "min_criteria": self.min_criteria,
            "enabled_keys": list(self.enabled_keys),
        }


class MutationAggregator:
    def __init__(self, config: Dict = None):
        self.config = config or load_config()
        self.client = HTTPClient(self.config)
        self.cache = Cache(self.config.get("cache", {}).get("path", CACHE_DB), 
                           timeout=self.config.get("cache", {}).get("timeout", 30))
        self.output_dir = self.config.get("output", {}).get("directory", "outputs")
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Configurer le parser de variants avec les seuils de splicing
        VariantParser.set_config(self.config)

    def _init_adapters(self, uid: str):
        sources_cfg = self.config.get("sources", {})
        adapters = {}
        if sources_cfg.get("uniprot", {}).get("enabled", True):
            adapters["uniprot"] = UniProtAdapter(self.client, self.cache, self.config, uid)
        if sources_cfg.get("clinvar", {}).get("enabled", True):
            adapters["clinvar"] = ClinVarAdapter(self.client, self.cache, self.config, uid)
        if sources_cfg.get("pharmgkb", {}).get("enabled", True):
            adapters["pharmgkb"] = PharmGKBAdapter(self.client, self.cache, self.config, uid)
        return adapters

    def _correlate_records(self, all_records: List[MutationRecord], run_logger) -> Tuple[Dict[str, MutationRecord], List[MutationRecord]]:
        """Corrèle tous les records de toutes les sources via le moteur multi-clé."""
        engine = CorrelationEngine(self.config)
        global_enrichments = []
        
        for rec in all_records:
            # Identifier les records d'enrichissement global (sans position/variant)
            if not rec.position_aa and not rec.rsid and not rec.hgvs_c and not rec.chr_position:
                if rec.description or rec.impact_category:
                    global_enrichments.append(rec)
                continue
            engine.add_record(rec)
        
        stats = engine.get_correlation_stats()
        run_logger.info("Correlation: %d total, %d merged, %d unique, %d multi-source (min_criteria=%d)", 
                       stats["total"], stats["merged"], stats["unique_records"], 
                       stats["multi_source_records"], stats["min_criteria"])
        
        return {r.key: r for r in engine.get_records()}, global_enrichments

    def analyze(self, uniprot_id: str, write: bool = True) -> pd.DataFrame:
        start_time = time.perf_counter()
        run_logger = logger.getChild(uniprot_id)
        run_logger.info("Starting analysis...")
        adapters = self._init_adapters(uniprot_id)
        
        # 1. UniProt (Base) - récupère aussi les identifiants du gène
        uni_records = []
        gene_info = {"symbol": None, "ncbi_gene_id": None}
        if "uniprot" in adapters:
            uni_records = adapters["uniprot"].fetch_data(uniprot_id)
            gene_info = adapters["uniprot"].get_gene_info(uniprot_id)
        run_logger.info("Gene: %s (NCBI ID: %s)", 
                       gene_info.get("symbol") or "Unknown",
                       gene_info.get("ncbi_gene_id") or "N/A")

        # 2. External (ClinVar, PharmGKB)
        # Recherche par identifiants stables: UniProt ID, NCBI Gene ID, rsIDs
        known_rsids = [r.rsid for r in uni_records if r.rsid]
        ncbi_gene_id = gene_info.get("ncbi_gene_id")
        
        external_data = {}
        max_w = self.config.get("limits", {}).get("max_workers", 4)
        worker_timeout = self.config.get("limits", {}).get("worker_timeout", 120)
        with ThreadPoolExecutor(max_workers=max_w) as ex:
            futures = {}
            if "clinvar" in adapters:
                # ClinVar: recherche par NCBI Gene ID + rsIDs connus
                futures[ex.submit(adapters["clinvar"].fetch_data, uniprot_id, ncbi_gene_id, known_rsids)] = "clinvar"
            if "pharmgkb" in adapters:
                # PharmGKB: recherche par UniProt ID (cross-reference)
                futures[ex.submit(adapters["pharmgkb"].fetch_data, uniprot_id, ncbi_gene_id)] = "pharmgkb"
            
            for fut in as_completed(futures):
                name = futures[fut]
                try:
                    # Timeout configurable pour éviter blocage infini sur une source
                    if worker_timeout and worker_timeout > 0:
                        external_data[name] = fut.result(timeout=worker_timeout)
                    else:
                        external_data[name] = fut.result()
                except Exception as e:
                    if isinstance(e, TimeoutError):
                        run_logger.error("Timeout (%ss) in %s, cancelling task", worker_timeout, name)
                        try: fut.cancel()
                        except: pass
                    else:
                        run_logger.error("Error in %s: %s", name, e)
                    external_data[name] = []

        # 3. Corrélation Multi-Sources avec Index Multi-Clé
        all_records = uni_records + external_data.get("clinvar", []) + external_data.get("pharmgkb", [])
        indexed, global_enrichments = self._correlate_records(all_records, run_logger)

        # Application Enrichissement Global (Contexte Seul, pas de PMIDs)
        for rec in indexed.values():
            for enrichment in global_enrichments:
                if not rec.impact_category and enrichment.impact_category:
                    rec.impact_category = enrichment.impact_category
                if not rec.description and enrichment.description:
                    rec.description = enrichment.description

        # 4. Export
        df = pd.DataFrame([r.to_dict() for r in indexed.values()]).fillna("")
        if write:
            prefix = os.path.join(self.output_dir, f"{uniprot_id}_mutations")
            # Excel
            df.to_excel(f"{prefix}.xlsx", index=False, engine="openpyxl")
            wb = load_workbook(f"{prefix}.xlsx")
            ws = wb.active
            # Style Header
            header_font = Font(bold=True, color="FFFFFF")
            fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = fill
            # Hyperlinks
            cols = {c.value: c.column for c in ws[1] if c.value}
            for row in range(2, ws.max_row + 1):
                # PubMed et ClinVar links
                for col_name, url_type in [("pubmed_ids", "pubmed"), ("clinvar_id", "clinvar")]:
                    if col_name in cols:
                        cell = ws.cell(row=row, column=cols[col_name])
                        ids = [i.strip() for i in str(cell.value).split(",") if i.strip()]
                        if ids:
                            if url_type == "pubmed":
                                cell.hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/?term={quote_plus(' OR '.join(ids))}"
                            else:
                                cell.hyperlink = f"https://www.ncbi.nlm.nih.gov/clinvar/?term={quote_plus(' OR '.join(ids))}"
                            cell.style = "Hyperlink"
                # dbSNP link pour rsID
                if "rsid" in cols:
                    cell = ws.cell(row=row, column=cols["rsid"])
                    rsid = str(cell.value).strip()
                    if rsid and rsid.startswith("rs"):
                        cell.hyperlink = f"https://www.ncbi.nlm.nih.gov/snp/{rsid}"
                        cell.style = "Hyperlink"
            wb.save(f"{prefix}.xlsx")
            
            # JSON
            with open(f"{prefix}.json", "w") as f:
                json.dump({"metadata": {"timestamp": datetime.now(UTC).isoformat(), "uniprot_id": uniprot_id}, 
                           "records": df.to_dict(orient="records")}, f, indent=2, default=str)
            
            duration = time.perf_counter() - start_time
            logger.info("Output written for %s (%.2fs)", uniprot_id, duration)
            
        return df

def load_config(path: str = CONFIG_PATH) -> Dict:
    if os.path.exists(path):
        try:
            with open(path) as f:
                return yaml.safe_load(f) or {}
        except: pass
    return {}

def main():
    parser = argparse.ArgumentParser(description="Mutation Aggregator v6 (Clean Evidence)")
    parser.add_argument("uniprot_id", nargs="?", help="UniProt ID")
    parser.add_argument("--batch", help="CSV file with uniprot_id column", default=None)
    parser.add_argument("--config", default=CONFIG_PATH)
    parser.add_argument("--no-write", action="store_true")
    args = parser.parse_args()

    if not args.uniprot_id and not args.batch:
        if os.path.exists(DEFAULT_BATCH_FILE):
            args.batch = DEFAULT_BATCH_FILE
        else:
            parser.error("Provide ID or batch file")

    agg = MutationAggregator(load_config(args.config))

    if args.batch:
        batch_df = pd.read_csv(args.batch)
        col = "UniProt ID" if "UniProt ID" in batch_df.columns else "uniprot_id"
        uids = batch_df[col].dropna().unique()
        for uid in tqdm(uids, desc="Processing"):
            try:
                agg.analyze(uid, write=not args.no_write)
            except Exception as e:
                logger.error("Failed %s: %s", uid, e)
    else:
        df = agg.analyze(args.uniprot_id, write=not args.no_write)
        if "mutation" in df: df = df.copy(); df["mutation"] = df["mutation"].str.replace("→", "->")
        print("\n" + df.to_string(index=False))

if __name__ == "__main__":
    main()
